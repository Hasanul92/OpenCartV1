import openpyxl
from openpyxl.styles import PatternFill     # we have imported this to fill the color somewhere if required

# file=r"C:\Users\Me\Desktop\caldata.xlsx"

def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.cell(rownum,columnno).value)

def writeData(file,sheetName,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(rownum,columnno).value=data
    workbook.save(file)

def fillGreenColor(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    sheet.cell(rownum,columnno).fill=fillGreenColor()
    workbook.save(file)

def fillRedColor(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    PatternFill(start_color='ff000',end_color='ff000',fill_type='solid')
    sheet.cell(rownum,columnno).fill=fillRedColor()
    workbook.save(file)
